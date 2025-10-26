import os
import io
from datetime import datetime
from sqlalchemy.orm import Session
from models import Candidate, InterviewAnswer
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import tempfile

def get_candidate_report_data(db: Session, candidate_id: int):
    """Получает данные кандидата для отчёта"""
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        return None
    
    # Получаем все ответы кандидата
    answers = db.query(InterviewAnswer).filter(
        InterviewAnswer.candidate_id == candidate_id,
        InterviewAnswer.is_valid == True
    ).all()
    
    # Группируем ответы по процессам
    processes_data = {}
    for answer in answers:
        process = answer.process
        if process not in processes_data:
            processes_data[process] = []
        processes_data[process].append(answer)
    
    # Сортируем ответы по номеру вопроса
    for process in processes_data:
        processes_data[process].sort(key=lambda x: x.question_number)
    
    return {
        'candidate': candidate,
        'processes_data': processes_data,
        'answers': answers
    }

def calculate_process_metrics(answers):
    """Рассчитывает метрики процесса"""
    iteration_time = 0
    frequency = 0
    session_count = 0
    tools = []
    
    for answer in answers:
        if answer.question_number == 2:  # Время итерации
            import re
            numbers = re.findall(r'\d+', answer.answer)
            if numbers:
                iteration_time = float(numbers[0])
        elif answer.question_number == 3:  # Частота
            import re
            numbers = re.findall(r'\d+', answer.answer)
            if numbers:
                frequency = float(numbers[0])
        elif answer.question_number == 4:  # Количество повторов
            import re
            numbers = re.findall(r'\d+', answer.answer)
            if numbers:
                session_count = float(numbers[0])
        elif answer.question_number == 6:  # Инструменты
            tools.append(answer.answer)
    
    total_time = iteration_time * frequency * session_count
    rate_per_minute = float(os.getenv("PROCESS_RATE_PER_MINUTE", "0.5"))
    process_cost = total_time * rate_per_minute
    
    return {
        'iteration_time': iteration_time,
        'frequency': frequency,
        'session_count': session_count,
        'total_time': total_time,
        'process_cost': process_cost,
        'tools': tools
    }

def generate_pdf_report(report_data, output_path):
    """Генерирует PDF отчёт"""
    doc = SimpleDocTemplate(output_path, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    # Заголовок
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1,  # Center
        textColor=colors.darkblue
    )
    
    story.append(Paragraph("ОТЧЁТ ПО ПРОЦЕССАМ СОТРУДНИКА", title_style))
    story.append(Spacer(1, 20))
    
    # Информация о кандидате
    candidate = report_data['candidate']
    story.append(Paragraph(f"<b>ФИО:</b> {candidate.full_name}", styles['Normal']))
    story.append(Paragraph(f"<b>Процессы:</b> {candidate.processes or 'Не указаны'}", styles['Normal']))
    story.append(Paragraph(f"<b>Дата формирования:</b> {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Процессы
    processes_data = report_data['processes_data']
    if processes_data:
        for process_name, answers in processes_data.items():
            story.append(Paragraph(f"<b>Процесс: {process_name}</b>", styles['Heading2']))
            
            # Метрики процесса
            metrics = calculate_process_metrics(answers)
            
            # Таблица с данными процесса
            process_data = [
                ['Параметр', 'Значение'],
                ['Время одной итерации (мин)', f"{metrics['iteration_time']}"],
                ['Частота выполнения', f"{metrics['frequency']}"],
                ['Количество повторов за сессию', f"{metrics['session_count']}"],
                ['Общее время (мин)', f"{metrics['total_time']:.1f}"],
                ['Стоимость процесса (₽)', f"{metrics['process_cost']:.2f}"],
                ['Инструменты', ', '.join(metrics['tools']) if metrics['tools'] else 'Не указаны']
            ]
            
            process_table = Table(process_data, colWidths=[3*inch, 3*inch])
            process_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(process_table)
            story.append(Spacer(1, 15))
            
            # Ответы на вопросы
            story.append(Paragraph("<b>Ответы на вопросы:</b>", styles['Heading3']))
            for answer in answers:
                story.append(Paragraph(f"<b>Вопрос {answer.question_number}:</b> {answer.question}", styles['Normal']))
                story.append(Paragraph(f"<b>Ответ:</b> {answer.answer}", styles['Normal']))
                story.append(Spacer(1, 10))
            
            story.append(Spacer(1, 20))
    
    # Подпись
    story.append(Spacer(1, 30))
    story.append(Paragraph("Сформировано автоматически системой DeepInterview", styles['Normal']))
    
    doc.build(story)

def generate_excel_report(report_data, output_path):
    """Генерирует Excel отчёт"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт по процессам"
    
    # Стили
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=12)
    normal_font = Font(size=11)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Заголовок
    ws['A1'] = "ОТЧЁТ ПО ПРОЦЕССАМ СОТРУДНИКА"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    
    # Информация о кандидате
    candidate = report_data['candidate']
    ws['A3'] = "ФИО:"
    ws['B3'] = candidate.full_name
    ws['A4'] = "Процессы:"
    ws['B4'] = candidate.processes or 'Не указаны'
    ws['A5'] = "Дата формирования:"
    ws['B5'] = datetime.now().strftime('%d.%m.%Y %H:%M')
    
    # Процессы
    processes_data = report_data['processes_data']
    if processes_data:
        row = 7
        for process_name, answers in processes_data.items():
            ws[f'A{row}'] = f"Процесс: {process_name}"
            ws[f'A{row}'].font = header_font
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            # Метрики процесса
            metrics = calculate_process_metrics(answers)
            
            # Таблица с данными процесса
            process_data = [
                ['Параметр', 'Значение', '', ''],
                ['Время одной итерации (мин)', f"{metrics['iteration_time']}", '', ''],
                ['Частота выполнения', f"{metrics['frequency']}", '', ''],
                ['Количество повторов за сессию', f"{metrics['session_count']}", '', ''],
                ['Общее время (мин)', f"{metrics['total_time']:.1f}", '', ''],
                ['Стоимость процесса (₽)', f"{metrics['process_cost']:.2f}", '', ''],
                ['Инструменты', ', '.join(metrics['tools']) if metrics['tools'] else 'Не указаны', '', '']
            ]
            
            for i, (param, value, _, _) in enumerate(process_data):
                ws[f'A{row}'] = param
                ws[f'B{row}'] = value
                if i == 0:  # Заголовок
                    ws[f'A{row}'].fill = header_fill
                    ws[f'B{row}'].fill = header_fill
                    ws[f'A{row}'].font = header_font
                    ws[f'B{row}'].font = header_font
                else:
                    ws[f'A{row}'].font = normal_font
                    ws[f'B{row}'].font = normal_font
                row += 1
            
            # Ответы на вопросы
            ws[f'A{row}'] = "Ответы на вопросы:"
            ws[f'A{row}'].font = header_font
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            for answer in answers:
                ws[f'A{row}'] = f"Вопрос {answer.question_number}:"
                ws[f'A{row}'].font = header_font
                ws[f'B{row}'] = answer.question
                ws[f'B{row}'].font = normal_font
                row += 1
                ws[f'A{row}'] = "Ответ:"
                ws[f'A{row}'].font = header_font
                ws[f'B{row}'] = answer.answer
                ws[f'B{row}'].font = normal_font
                row += 2
            
            row += 1
    
    # Подпись
    ws[f'A{row}'] = "Сформировано автоматически системой DeepInterview"
    ws[f'A{row}'].font = normal_font
    
    # Автоподбор ширины колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)

def generate_report_files(db: Session, candidate_id: int):
    """Генерирует файлы отчёта и возвращает пути к ним"""
    report_data = get_candidate_report_data(db, candidate_id)
    if not report_data:
        return None, None
    
    # Создаём временные файлы
    pdf_path = tempfile.mktemp(suffix='.pdf')
    excel_path = tempfile.mktemp(suffix='.xlsx')
    
    try:
        generate_pdf_report(report_data, pdf_path)
        generate_excel_report(report_data, excel_path)
        return pdf_path, excel_path
    except Exception as e:
        # Очищаем файлы в случае ошибки
        for path in [pdf_path, excel_path]:
            if os.path.exists(path):
                os.remove(path)
        raise e

